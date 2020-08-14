import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")

print(wb.sheetnames)

sheet = wb["Sheet1"]

#wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

cell = sheet["a1"]
#print(cell.value)
#print(cell.row)

cell = sheet.cell(row = 1, column = 1)
